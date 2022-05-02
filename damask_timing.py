from functools import total_ordering
import random
import damask, time
from openpyxl import Workbook, workbook, load_workbook

workbook = load_workbook(filename="metingen WV py-rot-fast.xlsx")
workbook.sheetnames
['Sheet 1']
sheet = workbook.active

#time
def operationEu(a, l):
    totaltime = 0
    for i in range(l):
        processST = time.process_time()
        b = a.as_Euler_angles()
        c = damask.Rotation.from_Euler_angles(b)
        processET = time.process_time()
        totaltime += processET - processST
    print(totaltime/l * 10**3)
    return(totaltime/l* 10**3)

def operationOm(a, l):
    totaltime = 0
    for i in range(l):
        processST = time.process_time()
        b = a.as_matrix()
        c = damask.Rotation.from_matrix(b)
        processET = time.process_time()
        totaltime += processET - processST
    print(totaltime/l * 10**3)
    return(totaltime/l* 10**3)

def operationAx(a, l):
    totaltime = 0
    for i in range(l):
        processST = time.process_time()
        b = a.as_axis_angle()
        c = damask.Rotation.from_axis_angle(b)
        processET = time.process_time()
        totaltime += processET - processST
    print(totaltime/l * 10**3)
    return(totaltime/l* 10**3)

def operationHo(a, l):
    totaltime = 0
    for i in range(l):
        processST = time.process_time()
        b = a.as_homochoric()
        c = damask.Rotation.from_homochoric(b)
        processET = time.process_time()
        totaltime += processET - processST
    print(totaltime/l * 10**3)
    return(totaltime/l* 10**3)

def operationRo(a, l):
    totaltime = 0
    for i in range(l):
        processST = time.process_time()
        b = a.as_Rodrigues_vector()
        c = damask.Rotation.from_Rodrigues_vector(b)
        processET = time.process_time()
        totaltime += processET - processST
    print(totaltime/l * 10**3)
    return(totaltime/l* 10**3)

#a zijn de rotaties, l is aantal herhalingen van de test
def oneTest(a, l, row):
    row = str(row)
    result = operationEu(a, l)
    sheet["A"+row] = str(result)
    result = operationOm(a, l)
    sheet["B"+row] = str(result)
    result = operationAx(a, l)
    sheet["C"+row] = str(result)
    result = operationHo(a, l)
    sheet["D"+row] = str(result)
    result = operationRo(a, l)
    sheet["E"+row] = str(result)

def longTestRun():
    print("EU - OM - AX - HO - RO, all times in ms")
    min = 1
    max = 10
    for i in range (min,max):
        sheet["F"+str(2+i-min)] = 2**i
        l = 1000
        n = 2**i
        print(n)
        a = damask.Rotation.from_random(n)
        oneTest(a, l, 2+i-min)

def main():
    sheet["A1"] = "EU"
    sheet["B1"] = "OM"
    sheet["C1"] = "AX"
    sheet["D1"] = "HO"
    sheet["E1"] = "RO"
    sheet["F1"] = "n"
    sheet["G1"] = "all times are in ms"
main()
longTestRun()
rand = random.randrange(1, 100000)
workbook.save("meting WV py: "+str(rand)+".xlsx")
