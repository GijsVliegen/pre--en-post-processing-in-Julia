
from openpyxl import Workbook, workbook, load_workbook
import random
workbook = load_workbook(filename="meting WV jl.xlsx")
workbook.sheetnames
['Sheet 1']
sheet = workbook.active

with open('/home/gijs/free/programmeren/julia/WV/rotations.jl/juliaTimingResults', 'r') as file:
    data = file.readlines()

def main():
    sheet["A1"] = "OM"
    sheet["B1"] = "EU"
    sheet["C1"] = "AX"
    sheet["D1"] = "HO"
    sheet["E1"] = "RO"
    sheet["F1"] = "n"
    sheet["G1"] = "NOT all times are in ms"
    sheet["H1"] = "eenheid om"
    sheet["I1"] = "eenheid eu"
    sheet["J1"] = "eenheid ax"
    sheet["K1"] = "eenheid ho"
    sheet["L1"] = "eenheid ro"
#altijd eerst om, dan eu, dan ax, dan ho, dan ro
main()

rowCounter = 2
columnCounter = 0
columnString = "ABCDEF"
columnStringEenheid = "HIJKL"

for i in data:
    print(i)
    if i.startswith("voor"):
        n = i[9:-1] #nakijken
        sheet[columnString[5]+str(rowCounter)] = str(n)
    if i.startswith("  "):
        end = i.find("(") - 4 #nakijken
        eenheid = i[end+1:end+3]
        if eenheid.startswith(" "):
            res = i[2: end+1]
        else:
            res = i[2: end]
        print("res = ", res)
        print("eenheid = ", eenheid)
        sheet[columnString[columnCounter]+str(rowCounter)] = str(res)
        sheet[columnStringEenheid[columnCounter]+str(rowCounter)] = str(eenheid)
        columnCounter += 1
    if columnCounter == 5:
        columnCounter = 0
        rowCounter += 1

rand = random.randrange(1, 100000)
workbook.save("meting WV jl.xlsx")